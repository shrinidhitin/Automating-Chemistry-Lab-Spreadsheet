import openpyxl as xl

class LinearRegression:
    def __init__(self, x, y):
        self.x = x
        self.y = y
        sum_x = sum(self.x)
        sum_y = sum(self.y)
        sum_xy = sum(a*b for a, b in zip(self.x, self.y))
        sum_x2 = sum(a**2 for a in self.x)
        n = len(self.x)
        self.m = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x ** 2)
        self.b = (sum_y * sum_x2 - sum_x * sum_xy) / (n * sum_x2 - sum_x ** 2)


    def uncertainty(self):
        n = len(self.x)
        residuals = [self.y[i] - (self.m * self.x[i] + self.b) for i in range(n)]
        residual_sum_of_squares = sum(r ** 2 for r in residuals)
        variance = residual_sum_of_squares / (n - 2)
        x_mean = sum(self.x) / n
        s_xx = sum((x_i - x_mean) ** 2 for x_i in self.x)
        m_uncertainty = (variance / s_xx) ** 0.5
        b_uncertainty = (variance * (1/n + x_mean**2 / s_xx)) ** 0.5
        return m_uncertainty, b_uncertainty
    

    def line_equation(self):
        return self.m, self.b

wb = xl.load_workbook('Chem Lab Data - Copy.xlsx')
sheet = wb['Sheet1']
output1_row = 16
output2_row = 25
current = 0.5
current_uncertainty = 0.05
molar_mass = 63.55
F = 96485.33

for row in range(6, 11):
    cell_1 = sheet.cell(row, 4) 
    cell_2 = sheet.cell(row, 5)
    cell_3 = sheet.cell(row, 6)
    cell_4 = sheet.cell(row, 7)
    cell_5 = sheet.cell(row, 8)
    cell_6 = sheet.cell(row, 9)
    t1_change_in_mass = float(cell_2.value) - float(cell_1.value)
    t2_change_in_mass = float(cell_4.value) - float(cell_3.value)
    t3_change_in_mass = float(cell_6.value) - float(cell_5.value)

    sheet.cell(output1_row, 5).value = t1_change_in_mass
    sheet.cell(output1_row, 6).value = t2_change_in_mass
    sheet.cell(output1_row, 7).value = t3_change_in_mass
    output1_row += 1

for row in range(16, 21):
    time = sheet.cell(row, 4).value
    cell_1 = sheet.cell(row, 5) 
    cell_2 = sheet.cell(row, 6)
    cell_3 = sheet.cell(row, 7)
    average_change_in_mass = (float(cell_1.value) + float(cell_2.value) + float(cell_3.value)) / 3
    sheet.cell(row, 8).value = average_change_in_mass
    sheet.cell(output2_row, 7).value = average_change_in_mass

    theoretical_change_in_mass = (time * current * molar_mass)/(2 * F)
    sheet.cell(output2_row, 5).value = theoretical_change_in_mass

    uncertainty_theoretical_change_in_mass = (time * current_uncertainty * molar_mass)/(2 * F)
    sheet.cell(output2_row, 6).value = uncertainty_theoretical_change_in_mass

    P_uncertainty_change_in_mass = (average_change_in_mass / theoretical_change_in_mass) * 100
    sheet.cell(output2_row, 8).value = P_uncertainty_change_in_mass

    PE_of_trials = ((average_change_in_mass - theoretical_change_in_mass) / theoretical_change_in_mass) * 100
    sheet.cell(output2_row, 9).value = PE_of_trials

    output2_row += 1

lr = LinearRegression([sheet.cell(row, 4).value for row in range(16, 21)], [sheet.cell(row, 8).value for row in range(16, 21)])
m_uncertainty, b_uncertainty = lr.uncertainty()
sheet["L7"] = lr.m
sheet["N7"] = m_uncertainty
sheet["M7"] = lr.b
sheet["O7"] = b_uncertainty
m, b = lr.line_equation()
sheet["L8"] = f"{m}x + {b}"

wb.save('Chem Lab Data - Copy2.xlsx')